#! python3
# ExcelColumnComparator.py - Tool designed for VLOOKUP operations in Excel files. It ignores case, size, and file extension. It saves time and reduces strain on your wrists.
# Author - Robert Stokłosa

import openpyxl, re
import os
from openpyxl.utils import get_column_letter, column_index_from_string
import time

# Welcome instruction
print(f'Excel Column Comparator - tool designed for VLOOKUP operations in Excel files.'
      '\nPlace the Excel file to be checked in the folder where this program is located.\n')

# Opening Excel file
while True:
    fileName = input('Write the name of the file you want to check: ')
    try:
        file = openpyxl.load_workbook(fileName + '.xlsx')
        break
    except Exception:
        try:
            file = openpyxl.load_workbook(fileName + '.xltm')
            break
        except Exception:
            try:
                file = openpyxl.load_workbook(fileName + '.xlsm')
                break
            except Exception:
                try:
                    file = openpyxl.load_workbook(fileName + '.xltx')
                    break
                except Exception:
                    print('The file you provided does not exist or its name is incorrect'
                          '\nPlease try again')
                continue
            
# Choosing sheet
while True:
    try:
        fileSheet = input('Write the name of the sheet you want to search in: ')
        sheet = file[fileSheet]
    except Exception:
        print('The sheet you provided does not exist or its name is incorrect'
            '\nPlease try again')
        continue
    else:
        break

# Inserting column with the file names
nameColumn = column_index_from_string(input('Write the letter of the column with the file names: '))
# Getting range of rows with filenames
for t1 in range(1, sheet.max_row+1):
    if sheet.cell(row=t1, column=nameColumn).value is not None:
         break
    
for l1 in range(t1, sheet.max_row+1):
    if sheet.cell(row=l1, column=nameColumn).value is None:
         break

# Inserting column to scrape
scrapeColumn = column_index_from_string(input('Write the letter of the column where you want to check for the presence of names: '))
# Getting range of rows to scrape
for t2 in range(1, sheet.max_row+1):
    if sheet.cell(row=t2, column=scrapeColumn).value is not None:
         break
    
for l2 in range(t2, sheet.max_row+1):
    if sheet.cell(row=l2, column=scrapeColumn).value is None:
         break
    
# Record start time
start = time.time()
print(f'\nThe program is comparing columns'
'\nPlease wait for the result...')

# Scraping file with column
for x in range(t1+1, l1):
    i = 0
    
    # Regex for file extension
    input_text = str(sheet.cell(row=x, column=nameColumn).value)
    regex = r"\.[a-zA-Z]{2,5}$"
    noExtension = re.sub(regex, '', input_text)

    for z in range(t2+1, l2):
        if noExtension.casefold() == str(sheet.cell(row=z, column=scrapeColumn).value).casefold():
            i += 1   
    
    # Filling a column with Yes or No based on the scraping results
    if i > 0:
        sheet['E%s' %(x)] = 'Yes [%s]' %(str(i)) # The program counts the numbers of phrases in column and display it in parenthesis
    elif i == 0:
        sheet['E%s' %(x)] = 'No'

# Saving results to new xlsx file with '_Sorted' ending
file.save(fileName + '_Sorted.xlsx')

# Record end time
end = time.time()

# Print execution time of script
print(f'Program execution time:', time.strftime('%H:%M:%S', time.gmtime(end-start)))
print(f"I wish you a pleasant work!\n")

# Closing program after pressing enter
input("Press ENTER to exit the program")
